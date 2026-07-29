"""Tests run against a synthetic workbook that mimics Avantage's export layout,
so no measured data is needed in the repository."""

from __future__ import annotations

import io

import numpy as np
import pytest
from openpyxl import Workbook

from xpsfig import export, parser, plotting, style, tables


def make_workbook(with_fit: bool = True, with_chem_states: bool = True) -> io.BytesIO:
    """Reproduce Avantage's sheet layout: metadata rows, header at 15, data at 17."""
    wb = Workbook()
    wb.remove(wb.active)

    energy = np.round(np.arange(298.08, 277.0, -0.1), 2)

    def gauss(center: float, width: float, height: float) -> np.ndarray:
        return height * np.exp(-((energy - center) ** 2) / (2 * width ** 2))

    background = np.full_like(energy, 100.0)
    comp_a = background + gauss(284.5, 0.7, 2800)
    comp_b = background + gauss(288.9, 1.5, 400)
    envelope = background + (comp_a - background) + (comp_b - background)
    raw = envelope + np.random.default_rng(0).normal(0, 20, energy.size)

    ws = wb.create_sheet("C1s Scan")
    ws["H1"] = "Acquisition Parameters :"
    ws["H3"], ws["I3"] = "Parameter", " "
    ws["H4"], ws["I4"] = "Number of Scans", 10
    ws["H5"], ws["I5"] = "Source Gun Type", "Al K Alpha"
    ws["A9"] = r"E:\data\sample.DATA\C1s Scan.VGD"

    ws["A15"] = "Binding Energy (E)"
    ws["A16"], ws["C16"] = "eV", "Counts / s"
    columns = [("E", "C1s Scan A", comp_a), ("F", "C1s Scan B", comp_b),
               ("G", "Backgnd.", background), ("H", "Envelope", envelope),
               ("I", "Residuals", raw - envelope)]
    if not with_fit:
        columns = [("E", "Backgnd.", background)]
    for col, name, _ in columns:
        ws[f"{col}15"] = name

    for i, be in enumerate(energy):
        row = 17 + i
        ws[f"A{row}"] = float(be)
        ws[f"C{row}"] = float(raw[i])
        for col, _, values in columns:
            # Avantage writes a literal 0 in the first row of fitted channels
            ws[f"{col}{row}"] = 0.0 if i == 0 else float(values[i])

    pt = wb.create_sheet("Peak Table")
    pt["A1"] = "Peak Table : "
    headers = [" ", "Name ", "Start BE", "Peak BE", "End BE", "Height CPS", "FWHM eV",
               "Area (P) CPS.eV", "Area (N) TPP-2M", "Weight %", "Peak Type", "Q ",
               "SF ALTHERMO1", "TXFN ", "Backgnd ", "PP Height CPS", "PP Hgt (N) ",
               "PP At. % ", "Title ", "File Name "]
    for j, name in enumerate(headers, start=1):
        pt.cell(row=2, column=j, value=name)

    peak_rows = [("C1s", 284.50, 1.81, 2543.46, 6414.75, 12.74, 50.85, "C1s Scan")]
    if with_fit:
        peak_rows = [
            ("C1s Scan A", 284.52, 1.63, 2851.58, 5434.02, 10.51, 35.75, "C1s Scan"),
            ("C1s Scan B", 288.92, 3.50, 404.30, 1657.52, 3.22, 6.91, "C1s Scan"),
        ]
    for i, (name, be, fwhm, height, area, weight, at_pct, title) in enumerate(peak_rows):
        row = 3 + i
        pt.cell(row=row, column=2, value=name)
        pt.cell(row=row, column=4, value=be)
        pt.cell(row=row, column=6, value=height)
        pt.cell(row=row, column=7, value=fwhm)
        pt.cell(row=row, column=8, value=area)
        pt.cell(row=row, column=10, value=weight)
        pt.cell(row=row, column=18, value=at_pct)
        pt.cell(row=row, column=19, value=title)

    if with_chem_states:
        base = 3 + len(peak_rows) + 2
        pt.cell(row=base, column=1, value="Chemical State Assessment :")
        pt.cell(row=base + 1, column=1, value="Chemical Composition of Sample ")
        pt.cell(row=base + 1, column=2, value="Spectral Region ")
        pt.cell(row=base + 1, column=3, value="Details ")
        for k, (compound, region, detail) in enumerate([
            (" carbon", "C1s", "C-C or C-H"),
            (" carbonate", "C1s", "carbonate"),
            (None, "O1s", "Metal CO3"),
            ("??? carbide", "C1s", "carbide"),
        ]):
            row = base + 2 + k
            if compound:
                pt.cell(row=row, column=1, value=compound)
            pt.cell(row=row, column=2, value=region)
            pt.cell(row=row, column=3, value=detail)
        pt.cell(row=base + 6, column=1, value="??? = Lower confidence assignment")

    titles = wb.create_sheet("Titles")
    titles["A1"] = "15:35:32  Wednesday, January 29, 2025"
    titles["A3"] = "C:\\d.xlsx"
    titles["A5"] = "TestSample-decon1"
    titles["A7"] = "Data Files :"

    wb.create_sheet("Sheet1")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def dataset() -> parser.Dataset:
    return parser.load_workbook_dataset(make_workbook(), label="Test")


# --- parsing ---------------------------------------------------------------
def test_region_and_series_detected(dataset):
    assert list(dataset.regions) == ["C1s"]
    region = dataset.regions["C1s"]
    assert region.raw is not None
    assert [s.name for s in region.components] == ["C1s Scan A", "C1s Scan B"]
    assert len(region.by_kind(parser.BACKGROUND)) == 1
    assert len(region.by_kind(parser.ENVELOPE)) == 1
    assert region.has_fit


def test_energy_axis_is_descending_and_complete(dataset):
    energy = dataset.regions["C1s"].energy
    assert energy[0] > energy[-1]
    assert energy.size == 211
    assert dataset.regions["C1s"].energy_range == pytest.approx((277.08, 298.08), abs=0.01)


def test_zero_edge_artefact_is_masked(dataset):
    """Avantage's leading 0 in fitted channels must not survive parsing."""
    region = dataset.regions["C1s"]
    for series in region.components + region.by_kind(parser.BACKGROUND):
        assert np.isnan(series.y[0])
    assert not np.isnan(region.raw.y[0])


def test_zero_edge_kept_when_disabled():
    ds = parser.load_workbook_dataset(make_workbook(), drop_zeros=False)
    assert ds.regions["C1s"].components[0].y[0] == 0.0


def test_title_used_as_default_label():
    ds = parser.load_workbook_dataset(make_workbook())
    assert ds.title == "TestSample-decon1"
    assert ds.label == "TestSample-decon1"


def test_component_name_split():
    assert parser.split_component_name("La3d Scan A") == ("La3d", "A")
    assert parser.split_component_name("O1s") == ("O1s", None)
    assert parser.region_from_sheet("Zn2p Scan_0") == "Zn2p"
    assert parser.region_from_sheet("XPS Survey") == "Survey"


# --- peak table ------------------------------------------------------------
def test_peaks_parsed_with_source(dataset):
    assert [p.name for p in dataset.peaks] == ["C1s Scan A", "C1s Scan B"]
    assert all(p.source == "fit" for p in dataset.peaks)
    assert dataset.peaks[0].peak_be == 284.52
    assert dataset.peaks[0].atomic_pct == 35.75


def test_chemical_state_block_is_not_read_as_peaks(dataset):
    assert not any("Spectral" in p.name for p in dataset.peaks)
    assert len(dataset.peaks) == 2


def test_chemical_states_forward_fill_compound(dataset):
    states = dataset.chemical_states
    assert [(s.compound, s.region) for s in states][:3] == [
        ("carbon", "C1s"), ("carbonate", "C1s"), ("carbonate", "O1s"),
    ]
    assert states[-1].compound == "carbide"
    assert states[-1].confident is False


# --- merging ---------------------------------------------------------------
def test_merge_prefers_fitted_region_and_dedupes_peaks():
    fitted = parser.load_workbook_dataset(make_workbook(with_fit=True), label="s")
    plain = parser.load_workbook_dataset(make_workbook(with_fit=False), label="s")

    merged = parser.merge_datasets([plain, fitted], "s")
    assert merged.regions["C1s"].has_fit
    # merging the same file twice must not duplicate rows
    assert len(parser.merge_datasets([fitted, fitted], "s").peaks) == len(fitted.peaks)


# --- plotting --------------------------------------------------------------
def test_normalization_modes():
    y = np.array([1.0, 3.0, 5.0])
    assert plotting.normalize(y, "max").max() == pytest.approx(1.0)
    assert plotting.normalize(y, "minmax").tolist() == pytest.approx([0.0, 0.5, 1.0])
    assert plotting.normalize(y, "area").sum() == pytest.approx(1.0)
    assert plotting.normalize(y, "none").tolist() == y.tolist()


def test_stack_offsets_are_evenly_spaced():
    curves = [np.array([0.0, 1.0]), np.array([0.0, 0.5])]
    assert plotting.stack_offsets(curves, 0.5) == pytest.approx([0.0, 0.5])


def test_render_stack_and_decon_figure(dataset):
    style.apply_style("DejaVu Sans", 9)
    region = dataset.regions["C1s"]

    stack = plotting.Panel(
        kind="stack", letter="a", inner_label="C1s",
        curves=[plotting.Curve(x=region.energy, y=region.raw.y, label="S1", color="#000000")],
    )
    decon = plotting.Panel(
        kind="decon", letter="b", legend_mode="inside",
        decon=plotting.DeconCurves(
            x=region.energy, raw=region.raw.y,
            components=[(s.short_name, s.y, c) for s, c in
                        zip(region.components, style.fill_colors("Pastel (Avantage benzeri)", 2))],
            background=region.by_kind(parser.BACKGROUND)[0].y,
            envelope=region.by_kind(parser.ENVELOPE)[0].y,
            sample_label="Test",
        ),
    )
    spec = plotting.FigureSpec(panels=[stack, decon], nrows=1, ncols=2,
                              shared_legend=True, shared_legend_entries=[("S1", "#000000")])
    fig = plotting.render_figure(spec)

    assert len(fig.axes) == 2
    # binding energy axis must run high -> low
    assert fig.axes[0].get_xlim()[0] > fig.axes[0].get_xlim()[1]
    legend = fig.axes[1].get_legend()
    assert [t.get_text() for t in legend.get_texts()][0] == "Raw Data"


@pytest.mark.parametrize("fmt", ["PNG", "TIFF", "PDF", "SVG"])
def test_export_formats_produce_output(dataset, fmt):
    style.apply_style("DejaVu Sans", 9)
    region = dataset.regions["C1s"]
    panel = plotting.Panel(
        curves=[plotting.Curve(x=region.energy, y=region.raw.y, label="S", color="#000000")])
    fig = plotting.render_figure(plotting.FigureSpec(panels=[panel]))
    assert len(export.figure_bytes(fig, fmt, 300)) > 1000


def test_bundle_contains_every_combination(dataset):
    style.apply_style("DejaVu Sans", 9)
    region = dataset.regions["C1s"]
    panel = plotting.Panel(
        curves=[plotting.Curve(x=region.energy, y=region.raw.y, label="S", color="#000000")])
    fig = plotting.render_figure(plotting.FigureSpec(panels=[panel]))

    import zipfile
    data = export.bundle(fig, "fig", ["PNG", "TIFF", "PDF"], [300, 600])
    names = zipfile.ZipFile(io.BytesIO(data)).namelist()
    assert sorted(names) == [
        "fig.pdf", "fig_300dpi.png", "fig_300dpi.tiff", "fig_600dpi.png", "fig_600dpi.tiff",
    ]


# --- tables ----------------------------------------------------------------
def test_summary_table_layout(dataset):
    df = tables.summary_table([dataset], metrics=["peak_be", "weight_pct"],
                              labels=["850 °C"], source="fit")
    assert list(df.columns) == ["C1s"]
    assert df.index.tolist() == [("850 °C", "Peak BE (eV)"), ("850 °C", "Weight (%)")]
    # "main" picks the most intense component
    assert df.loc[("850 °C", "Peak BE (eV)"), "C1s"] == "284.52"


def test_summary_table_prefers_requested_source():
    fit = parser.load_workbook_dataset(make_workbook(with_fit=True), label="s")
    core = parser.load_workbook_dataset(make_workbook(with_fit=False), label="s")
    merged = parser.merge_datasets([core, fit], "s")

    core_df = tables.summary_table([merged], metrics=["peak_be"], source="core")
    fit_df = tables.summary_table([merged], metrics=["peak_be"], source="fit")
    assert core_df.iloc[0]["C1s"] == "284.50"
    assert fit_df.iloc[0]["C1s"] == "284.52"


def test_element_order_falls_back_when_source_missing(dataset):
    # dataset only has fitted rows; asking for "core" must still list C1s
    assert tables.element_order([dataset], "core") == ["C1s"]


def test_detailed_table_lists_each_component(dataset):
    df = tables.detailed_table([dataset], labels=["850 °C"])
    assert df["Component"].tolist() == ["A", "B"]
    assert df["FWHM (eV)"].tolist() == ["1.63", "3.50"]


def test_chemical_state_table(dataset):
    df = tables.chemical_state_table([dataset], labels=["S"])
    assert len(df) == 4
    assert df["Confidence"].tolist()[-1] == "low"


def test_flatten_blanks_repeated_sample_names(dataset):
    df = tables.summary_table([dataset], metrics=["peak_be", "weight_pct"], labels=["S"])
    header, rows = tables.flatten(df, merge_sample_column=True)
    assert header[:2] == ["Sample", "Parameter"]
    assert rows[0][0] == "S" and rows[1][0] == ""


def test_docx_and_xlsx_export(dataset):
    df = tables.summary_table([dataset], metrics=["peak_be"], labels=["S"])
    assert tables.to_docx(df, "Table 1").startswith(b"PK")
    assert tables.to_xlsx(df, caption="Table 1").startswith(b"PK")
    assert r"\begin{tabular}" in tables.to_latex(df, "Table 1")


def test_docx_rejects_empty_table():
    import pandas as pd
    with pytest.raises(ValueError):
        tables.to_docx(pd.DataFrame(), "x")
