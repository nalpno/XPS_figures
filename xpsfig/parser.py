"""Thermo Avantage .xlsx export parser.

Avantage writes one worksheet per acquired region plus two bookkeeping sheets
(``Peak Table`` and ``Titles``).  Every region sheet has the same shape::

    rows  1-14  acquisition metadata (columns H/I hold parameter/value pairs)
    row     15  series names   (A = "Binding Energy (E)", C = raw counts,
                                then one column per fitted component, followed
                                by "Backgnd.", "Envelope", "Residuals")
    row     16  units
    rows  17->  numeric data
    columns B and D are always empty spacers

Survey and un-deconvoluted core files use the same layout with fewer series
(raw counts in C, background in E).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# series kinds
# ---------------------------------------------------------------------------
RAW = "raw"
BACKGROUND = "background"
ENVELOPE = "envelope"
RESIDUAL = "residual"
COMPONENT = "component"

_META_SHEETS = {"peak table", "titles", "sheet1"}
_BE_HEADER = "binding energy"


def classify_series(name: str) -> str:
    low = name.strip().lower().rstrip(".")
    if low in {"raw data", "raw", "counts / s"}:
        return RAW
    if low.startswith("backgnd") or low.startswith("background"):
        return BACKGROUND
    if low.startswith("envelope"):
        return ENVELOPE
    if low.startswith("residual"):
        return RESIDUAL
    return COMPONENT


def region_from_sheet(sheet_name: str) -> str:
    """``"Zn2p Scan"`` -> ``"Zn2p"``; ``"XPS Survey"`` -> ``"Survey"``."""
    name = sheet_name.strip()
    if name.lower().startswith("xps survey") or name.lower() == "survey":
        return "Survey"
    name = re.sub(r"\s*Scan(_\d+)?\s*$", "", name, flags=re.IGNORECASE)
    return name.strip() or sheet_name.strip()


def split_component_name(name: str) -> tuple[str, str | None]:
    """``"La3d Scan A"`` -> ``("La3d", "A")``; ``"O1s"`` -> ``("O1s", None)``."""
    m = re.match(r"^\s*(.+?)\s+Scan(?:_\d+)?\s*(\S+)?\s*$", name, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip(), (m.group(2).strip() if m.group(2) else None)
    return name.strip(), None


# ---------------------------------------------------------------------------
# data containers
# ---------------------------------------------------------------------------
@dataclass
class Series:
    """One curve inside a region (raw data, a fitted component, ...)."""

    name: str
    kind: str
    y: np.ndarray

    @property
    def short_name(self) -> str:
        element, comp = split_component_name(self.name)
        return f"{element} {comp}" if comp else element


@dataclass
class Region:
    name: str                       # "Zn2p", "Survey", ...
    sheet: str                      # original worksheet title
    energy: np.ndarray              # binding energy axis (eV)
    series: list[Series] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def by_kind(self, kind: str) -> list[Series]:
        return [s for s in self.series if s.kind == kind]

    @property
    def raw(self) -> Series | None:
        found = self.by_kind(RAW)
        return found[0] if found else None

    @property
    def components(self) -> list[Series]:
        return self.by_kind(COMPONENT)

    @property
    def has_fit(self) -> bool:
        return bool(self.components)

    @property
    def energy_range(self) -> tuple[float, float]:
        return float(np.nanmin(self.energy)), float(np.nanmax(self.energy))


@dataclass
class PeakRow:
    """One row of the ``Peak Table`` worksheet."""

    name: str
    element: str
    component: str | None
    peak_be: float | None
    fwhm: float | None
    height: float | None
    area_p: float | None
    area_n: float | None
    weight_pct: float | None
    atomic_pct: float | None
    region: str
    source: str = "core"        # "survey" | "core" | "fit"

    @property
    def is_empty(self) -> bool:
        return all(
            getattr(self, f) is None
            for f in ("peak_be", "fwhm", "height", "area_p", "weight_pct", "atomic_pct")
        )

    @property
    def key(self) -> tuple:
        return (self.name, self.region, self.peak_be, self.area_p)


@dataclass
class ChemicalState:
    """A row of Avantage's "Chemical State Assessment" block."""

    compound: str
    region: str
    detail: str
    confident: bool = True


@dataclass
class Dataset:
    """Everything read from a single Avantage workbook."""

    label: str                      # user facing sample name
    source: str                     # original filename
    regions: dict[str, Region] = field(default_factory=dict)
    peaks: list[PeakRow] = field(default_factory=list)
    chemical_states: list[ChemicalState] = field(default_factory=list)
    title: str = ""                 # name Avantage stored in the Titles sheet

    @property
    def region_names(self) -> list[str]:
        return list(self.regions.keys())

    def peaks_for(self, element: str) -> list[PeakRow]:
        return [p for p in self.peaks if p.element.lower() == element.lower()]


# ---------------------------------------------------------------------------
# low level helpers
# ---------------------------------------------------------------------------
def _as_float(value: Any) -> float:
    if value is None:
        return float("nan")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return float("nan")


def _find_header_row(rows: list[tuple], limit: int = 40) -> int | None:
    for idx, row in enumerate(rows[:limit]):
        first = row[0] if row else None
        if isinstance(first, str) and first.strip().lower().startswith(_BE_HEADER):
            return idx
    return None


def _read_metadata(rows: list[tuple]) -> dict[str, Any]:
    """Acquisition parameters live in columns H/I of the first ~14 rows."""
    meta: dict[str, Any] = {}
    for row in rows[:16]:
        if len(row) < 9:
            continue
        key, value = row[7], row[8]
        if isinstance(key, str) and key.strip() and value not in (None, " "):
            k = key.strip()
            if k.lower() in {"parameter", "acquisition parameters :"}:
                continue
            meta[k] = value
    for row in rows[:12]:
        first = row[0] if row else None
        if isinstance(first, str) and ".VGD" in first.upper():
            meta["Source File"] = first.strip()
            break
    return meta


def _parse_region_sheet(sheet_name: str, rows: list[tuple], drop_zeros: bool) -> Region | None:
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return None

    headers = rows[header_idx]
    data_rows = [r for r in rows[header_idx + 2:] if r and r[0] is not None]
    if not data_rows:
        return None

    width = max(len(headers), max(len(r) for r in data_rows))
    table = np.full((len(data_rows), width), np.nan)
    for i, row in enumerate(data_rows):
        for j, value in enumerate(row):
            table[i, j] = _as_float(value)

    energy = table[:, 0]
    keep = ~np.isnan(energy)
    table, energy = table[keep], energy[keep]
    if energy.size == 0:
        return None

    region = Region(
        name=region_from_sheet(sheet_name),
        sheet=sheet_name.strip(),
        energy=energy,
        metadata=_read_metadata(rows),
    )

    for col in range(1, width):
        column = table[:, col]
        if np.all(np.isnan(column)):
            continue
        header = headers[col] if col < len(headers) else None
        name = str(header).strip() if isinstance(header, str) and header.strip() else None
        if name is None:
            # unnamed column holding numbers == the raw counts channel (column C)
            name = "Raw Data" if not region.by_kind(RAW) else f"Series {col}"
        kind = classify_series(name)
        values = column.copy()
        if drop_zeros and kind in {COMPONENT, BACKGROUND, ENVELOPE}:
            # Avantage writes a literal 0 at the very edge of fitted channels,
            # which would otherwise drop the curve to the axis floor.
            values[values == 0.0] = np.nan
        region.series.append(Series(name=name, kind=kind, y=values))

    return region if region.series else None


def _parse_peak_table(rows: list[tuple]) -> list[PeakRow]:
    header_idx = None
    for idx, row in enumerate(rows[:10]):
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip().lower() == "name":
            header_idx = idx
            break
    if header_idx is None:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[header_idx]]

    def col(*candidates: str) -> int | None:
        for cand in candidates:
            if cand in headers:
                return headers.index(cand)
        return None

    idx_name = col("name")
    idx_be = col("peak be")
    idx_fwhm = col("fwhm ev", "fwhm")
    idx_height = col("height cps")
    idx_area_p = col("area (p) cps.ev")
    idx_area_n = col("area (n) tpp-2m")
    idx_weight = col("weight %")
    idx_at = col("pp at. %", "at. %", "atomic %")
    idx_title = col("title")

    def get(row: tuple, index: int | None) -> float | None:
        if index is None or index >= len(row):
            return None
        value = _as_float(row[index])
        return None if np.isnan(value) else value

    peaks: list[PeakRow] = []
    for row in rows[header_idx + 1:]:
        if not row or idx_name is None or idx_name >= len(row):
            continue
        # Avantage appends a "Chemical State Assessment" block below the peaks;
        # its rows are the only ones that put text in column A.
        first = row[0]
        if isinstance(first, str) and first.strip():
            break
        raw_name = row[idx_name]
        if not isinstance(raw_name, str) or not raw_name.strip():
            continue
        name = raw_name.strip()
        element, component = split_component_name(name)
        title = row[idx_title] if idx_title is not None and idx_title < len(row) else None
        region = region_from_sheet(str(title)) if title else element
        peak = PeakRow(
            name=name,
            element=element,
            component=component,
            peak_be=get(row, idx_be),
            fwhm=get(row, idx_fwhm),
            height=get(row, idx_height),
            area_p=get(row, idx_area_p),
            area_n=get(row, idx_area_n),
            weight_pct=get(row, idx_weight),
            atomic_pct=get(row, idx_at),
            region=region,
            source="survey" if region == "Survey" else ("fit" if component else "core"),
        )
        if not peak.is_empty:
            peaks.append(peak)
    return peaks


def _parse_chemical_states(rows: list[tuple]) -> list[ChemicalState]:
    """Read the "Chemical State Assessment" block that follows the peak list."""
    start = None
    for idx, row in enumerate(rows):
        first = row[0] if row else None
        if isinstance(first, str) and first.strip().lower().startswith("chemical composition"):
            start = idx + 1
            break
    if start is None:
        return []

    states: list[ChemicalState] = []
    last_compound, last_confident = "", True
    for row in rows[start:]:
        if not row or len(row) < 3:
            continue
        compound = str(row[0]).strip() if isinstance(row[0], str) else ""
        region = str(row[1]).strip() if isinstance(row[1], str) else ""
        detail = str(row[2]).strip() if isinstance(row[2], str) else ""
        if not region or region.startswith("???") or "lower confidence" in compound.lower():
            continue
        if compound:
            # a compound name heads a group; the rows below it repeat for each region
            last_confident = not compound.startswith("???")
            last_compound = compound.lstrip("? ").strip()
        states.append(
            ChemicalState(
                compound=last_compound,
                region=region,
                detail=detail,
                confident=last_confident,
            )
        )
    return states


def _parse_titles(rows: list[tuple]) -> str:
    for row in rows[:8]:
        first = row[0] if row else None
        if isinstance(first, str):
            text = first.strip()
            if not text or text.endswith(".xlsx") or text.startswith("Data Files"):
                continue
            if re.match(r"^\d{1,2}:\d{2}", text):
                continue
            return text
    return ""


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------
def load_workbook_dataset(source: Any, label: str | None = None, drop_zeros: bool = True) -> Dataset:
    """Read an Avantage workbook (path or file-like) into a :class:`Dataset`."""
    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        filename = Path(getattr(source, "name", str(source))).name
        dataset = Dataset(label=label or Path(filename).stem, source=filename)

        for ws in wb.worksheets:
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            sheet_key = ws.title.strip().lower()
            if sheet_key == "peak table":
                dataset.peaks = _parse_peak_table(rows)
                dataset.chemical_states = _parse_chemical_states(rows)
            elif sheet_key == "titles":
                dataset.title = _parse_titles(rows)
            elif sheet_key in _META_SHEETS:
                continue
            else:
                region = _parse_region_sheet(ws.title, rows, drop_zeros)
                if region is not None:
                    dataset.regions[region.name] = region

        if label is None and dataset.title:
            dataset.label = dataset.title
        return dataset
    finally:
        wb.close()


def merge_datasets(datasets: list[Dataset], label: str) -> Dataset:
    """Fold several workbooks of the *same* sample into one dataset.

    Useful because Avantage exports survey, core and deconvolution results as
    separate files.  Regions that appear more than once keep the richer copy
    (the one that carries fitted components).
    """
    merged = Dataset(label=label, source=" + ".join(d.source for d in datasets))
    seen_peaks: set[tuple] = set()
    seen_states: set[tuple] = set()
    for ds in datasets:
        for name, region in ds.regions.items():
            current = merged.regions.get(name)
            if current is None or (region.has_fit and not current.has_fit):
                merged.regions[name] = region
        for peak in ds.peaks:
            if peak.key not in seen_peaks:
                seen_peaks.add(peak.key)
                merged.peaks.append(peak)
        for state in ds.chemical_states:
            key = (state.compound, state.region, state.detail)
            if key not in seen_states:
                seen_states.add(key)
                merged.chemical_states.append(state)
        if not merged.title:
            merged.title = ds.title
    return merged
